'''
Módulo para leitura de planilhas do Excel.
'''
from xlrd import open_workbook_xls
from openpyxl import load_workbook
from module.comparator import data_finder

def xls_r(file_path, word_list):
    '''Função para abertura e leitura de planilhas no formato .xls'''
    aux, temp = dict(), dict()
    data_frame = open_workbook_xls(file_path)
    for n_sheet in range(data_frame.nsheets):
        sheet = data_frame.sheet_by_index(n_sheet)
        key = sheet.name
        for row in range(sheet.nrows):
            for col in range(sheet.ncols):
                cell_name = str(chr(col + 97).upper()) + str(row + 1)
                cell = str(sheet.cell_value(row, col))
                aux = data_finder(word_list, cell, key, file_path.suffix, cname=cell_name)
                if aux:
                    temp[key].append(aux[key][0]) if key in temp else temp.update(aux)
    if len(temp) > 0:
        return temp

def xlsx_r(file_path, word_list):
    '''Função para abertura e leitura de planilhas no formato .xlsx'''
    aux, temp = dict(), dict()
    dataframe = load_workbook(file_path)
    for sheet in dataframe.worksheets:
        sheet.active_cell
        key = sheet.title
        for row in range(0, sheet.max_row):
            hlp = int()
            for col in sheet.iter_cols(1, sheet.max_column):
                hlp += 1
                cell_name = str(chr(hlp + 96).upper()) + str(row + 1)
                cell = str(col[row].value)
                aux = data_finder(word_list, cell, key, file_path.suffix, cname=cell_name)
                if aux:
                    temp[key].append(aux[key][0]) if key in temp else temp.update(aux)
    if len(temp) > 0:
        return temp

if __name__ == '__main__':
    pass
