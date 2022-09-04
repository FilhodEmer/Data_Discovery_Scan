from pandas import read_excel
from glob import glob
from openpyxl import load_workbook

for file in glob (r'c:\Users\emers\Desktop\Faculdade\TCC\Scan_DataDiscovery\teste\*.xlsx'):
    if '~$' not in file:
        dataframe = load_workbook(file)
        for x in dataframe.worksheets:
            print('Aba:', x.title)
            x.active_cell
            for row in range(0, x.max_row):
                aux = int()
                for col in x.iter_cols(1, x.max_column):
                    aux += 1
                    cell = str(col[row].value)
                    if 'dado' in cell:
                        print('Célula {}{}: {}'.format(chr(aux + 96).upper(), row + 1, cell))
            print()

'''for file in glob('*.xls'):
    df = xlrd.open_workbook_xls(file)
    for x in range(df.nsheets):
        sheet = df.sheet_by_index(x)
        for i in range(sheet.nrows):
            for j in range(sheet.ncols):
                if 'dado' in str(sheet.cell_value(i, j)):
                    print('Aba: {}'.format(sheet.name))
                    print('Célula: {}{}\n{}'.format(chr(j + 97).upper(), i+1, sheet.cell_value(i, j), end='\t'))'''

'''try:
    for file in glob('*.xlsx'):
        aux = dict()
        newdf = read_excel(file)
        print(newdf, '\n')
        coluna = newdf.columns.tolist()
        print('Coluna: ', coluna, len(coluna))
        for x in newdf.columns.values:
            print(newdf[x].tolist(), len(newdf[x].tolist()))
except FileNotFoundError:
    pass'''